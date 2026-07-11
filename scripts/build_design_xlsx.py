"""Build a formatted multi-sheet Architecture Design Document as .xlsx."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\vchaurasia\Downloads\Concall_Intelligence_Architecture_Design.xlsx"

# ---- palette ---------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LBLUE = "D9E1F2"
LBLUE2 = "EAF0FA"
GREEN = "375623"
LGREEN = "E2EFDA"
AMBER = "BF8F00"
LAMBER = "FFF2CC"
GREY = "44546A"
LGREY = "F2F2F2"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TITLE_F = Font(name="Calibri", size=20, bold=True, color=WHITE)
SUB_F = Font(name="Calibri", size=11, color=WHITE)
H1_F = Font(name="Calibri", size=14, bold=True, color=WHITE)
HDR_F = Font(name="Calibri", size=11, bold=True, color=WHITE)
BODY_F = Font(name="Calibri", size=11, color="1A1A1A")
BOLD_F = Font(name="Calibri", size=11, bold=True, color="1A1A1A")
MONO_F = Font(name="Consolas", size=10, color="1A1A1A")

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def section_banner(ws, row, text, span, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = H1_F
    c.fill = fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26
    return row + 1


def header_row(ws, row, headers, color=NAVY):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_F
        c.fill = fill(color)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22
    return row + 1


def data_row(ws, row, values, zebra=False, wrap=True, bold_first=False, mono=False):
    band = LGREY if zebra else WHITE
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = fill(band)
        c.border = BORDER
        c.alignment = WRAP if wrap else LEFT
        if mono:
            c.font = MONO_F
        elif bold_first and i == 1:
            c.font = BOLD_F
        else:
            c.font = BODY_F
    return row + 1


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===========================================================================
wb = Workbook()

# ---------- Sheet 1: Cover ----------
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False
set_widths(ws, [3, 26, 30, 30, 20, 3])
ws.merge_cells("B2:E2")
c = ws["B2"]
c.value = "Concall Intelligence Platform"
c.font = TITLE_F
c.fill = fill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 46

ws.merge_cells("B3:E3")
c = ws["B3"]
c.value = "Architecture Design Document (ADD)"
c.font = Font(size=14, bold=True, color=WHITE)
c.fill = fill(BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 28

ws.merge_cells("B4:E4")
c = ws["B4"]
c.value = ("Quarterly earnings-call intelligence & stock-ranking platform. Tracks management "
           "guidance -> delivery, scores companies vs a 10-section Business Analysis Template, "
           "computes PEG valuation, and ranks stocks. Decision-support only — not investment advice.")
c.font = Font(size=11, italic=True, color="404040")
c.alignment = WRAP_C
ws.row_dimensions[4].height = 60

meta = [
    ("Document version", "1.0"),
    ("Date", "2026-07-11"),
    ("Owner", "@vchaurasia_microsoft"),
    ("Project root", r"C:\Users\vchaurasia\concall-intel\\"),
    ("Runtime", "Python 3 · FastAPI · SQLAlchemy · SQLite · uvicorn 127.0.0.1:8010"),
    ("Status", "Implemented — working prototype, 23 companies live"),
]
r = 6
for k, v in meta:
    ws.cell(row=r, column=2, value=k).font = BOLD_F
    ws.cell(row=r, column=2).fill = fill(LBLUE)
    ws.cell(row=r, column=2).border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    cc = ws.cell(row=r, column=3, value=v)
    cc.font = BODY_F
    cc.border = BORDER
    cc.alignment = LEFT
    for col in (4, 5):
        ws.cell(row=r, column=col).border = BORDER
    ws.row_dimensions[r].height = 20
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws.cell(row=r, column=2, value="Worksheets in this workbook")
c.font = H1_F
c.fill = fill(GREY)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[r].height = 24
r += 1
toc = [
    ("1. Cover", "Title, metadata, table of contents"),
    ("2. Purpose & Scope", "Problem, goals, non-goals"),
    ("3. Principles", "Architectural principles & realization"),
    ("4. Logical Architecture", "5-layer stack (presentation -> persistence)"),
    ("5. Components", "Module-by-module responsibilities"),
    ("6. Data Model", "Entities, FK contract, temporal model"),
    ("7. Data Flows", "Ingestion + ranking-refresh flows"),
    ("8. Valuation (PEG)", "Section 9 formula & composite integration"),
    ("9. Scalability", "Prototype vs scale-out deployment view"),
    ("10. Decisions & Risks", "Trade-offs, risks/mitigations, future work"),
]
for name, desc in toc:
    ws.cell(row=r, column=2, value=name).font = BOLD_F
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=2).fill = fill(LBLUE2)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    cc = ws.cell(row=r, column=3, value=desc)
    cc.font = BODY_F
    cc.border = BORDER
    cc.alignment = LEFT
    for col in (4, 5):
        ws.cell(row=r, column=col).border = BORDER
    r += 1

# ---------- Sheet 2: Purpose & Scope ----------
ws = wb.create_sheet("Purpose & Scope")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 40, 70])
r = 1
r = section_banner(ws, r, "2. Purpose & Scope", 3, NAVY)
r += 0
ws.cell(row=r, column=2, value="Problem statement").font = BOLD_F
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=3)
c = ws.cell(row=r, column=3, value="Investors read dozens of concall transcripts per company across years but cannot easily "
           "answer: did management actually deliver on what they promised? Manual tracking of guidance -> outcome "
           "across 30 companies x 20 quarters is infeasible.")
c.font = BODY_F; c.alignment = WRAP
ws.cell(row=r, column=2).alignment = WRAP
ws.row_dimensions[r].height = 60
r += 2

r = section_banner(ws, r, "Goals", 3, GREEN)
r = header_row(ws, r, ["#", "Goal"], GREEN)
goals = [
    "Persist ALL transcripts and extracted facts so history is never lost.",
    "Auto-extract metrics, guidance and red flags from each concall.",
    "Build guidance -> outcome chains across quarters (FY24 -> FY25 -> FY26).",
    "Score each company 0-100 vs a 10-section template (incl. Section 9 PEG valuation).",
    "Rank all companies 1 -> N on a composite of business quality, CAGR, ROE and valuation.",
    "Ingest new quarters incrementally; update only the affected company's dashboard.",
    "Be horizontally scalable and provider-agnostic (LLM, extractor, DB all swappable).",
]
for i, g in enumerate(goals, 1):
    r = data_row(ws, r, [i, g], zebra=(i % 2 == 0))
r += 1

r = section_banner(ws, r, "Non-goals", 3, AMBER)
r = header_row(ws, r, ["#", "Out of scope"], AMBER)
non = [
    "Real-time price feeds / trading execution.",
    "Portfolio management / broker integration.",
    "Guaranteeing external market-data completeness (gaps are flagged, never fabricated).",
]
for i, g in enumerate(non, 1):
    r = data_row(ws, r, [i, g], zebra=(i % 2 == 0))

# ---------- Sheet 3: Principles ----------
ws = wb.create_sheet("Principles")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 30, 80])
r = 1
r = section_banner(ws, r, "3. Architectural Principles", 3, NAVY)
r = header_row(ws, r, ["#", "Principle", "How it is realized"])
prin = [
    ("Separation of concerns", "Ingestion, Extraction, Analysis, Guidance, Scoring, API, UI are independent modules under app/."),
    ("Provider abstraction", "LLM (stub/openai/azure/agent) and extractor (pymupdf/pdfplumber/azure_docintel) selected by config; swap without code changes."),
    ("Idempotent pipeline", "Re-running any transcript is safe; each stage checks TranscriptStatus before acting."),
    ("DB-backed memory", "Every fact is a row; the 'compare against history' feature is a query, not a cache."),
    ("Stateless API", "FastAPI app holds no session state -> trivially replicated behind a load balancer."),
    ("Storage portability", "SQLAlchemy ORM -> SQLite today, Postgres tomorrow, zero API-contract change."),
    ("Honest data", "Missing data flagged (*, N/A, source flags M/C/G/P) rather than guessed."),
]
for i, (p, h) in enumerate(prin, 1):
    r = data_row(ws, r, [i, p, h], zebra=(i % 2 == 0), bold_first=False)
    ws.cell(row=r-1, column=2).font = BOLD_F

# ---------- Sheet 4: Logical Architecture ----------
ws = wb.create_sheet("Logical Architecture")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 26, 40, 55])
r = 1
r = section_banner(ws, r, "4. Logical Architecture — 5-Layer Stack", 4, NAVY)
r = header_row(ws, r, ["Layer", "Name", "Modules", "Responsibility"])
layers = [
    ("L5", "Presentation", "app/static: index.html, app.js, style.css", "Dashboard/leaderboards, per-company Insights (10 sections + PEG card), ranking table."),
    ("L4", "API", "app/main.py, api/routes_*", "FastAPI REST: upload, insights, /fundamental-ranking, leaderboards. Stateless."),
    ("L3", "Domain / Service", "app/services/*, app/workers/pipeline.py", "ingestion -> extraction -> analysis -> guidance -> scoring, orchestrated per transcript."),
    ("L2", "Persistence", "app/db.py, app/models.py", "SQLAlchemy ORM over SQLite (concall.db); 10 entity tables."),
    ("L1", "External / Offline", "scripts/*.py, data/*.json", "Yahoo Finance (CAGR, PE), BSE/NSE/Screener transcripts; offline analytics -> DB."),
]
colors = [LBLUE, LBLUE2, LGREEN, LAMBER, LGREY]
for (lid, name, mods, resp), col in zip(layers, colors):
    for i, v in enumerate([lid, name, mods, resp], start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.fill = fill(col); c.border = BORDER; c.alignment = WRAP
        c.font = BOLD_F if i <= 2 else BODY_F
    ws.row_dimensions[r].height = 46
    r += 1
r += 1
r = section_banner(ws, r, "Request path (top-down) and data path (bottom-up)", 4, GREY)
flow = [
    "PDF upload -> API -> Service pipeline -> Persistence  (write path)",
    "Browser -> API -> Persistence -> JSON -> SPA render   (read path)",
    "Offline scripts -> Yahoo/BSE -> data/*.json -> persist -> Persistence  (enrichment path)",
]
for i, f in enumerate(flow, 1):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value=f)
    c.font = MONO_F; c.fill = fill(WHITE if i % 2 else LGREY); c.border = BORDER; c.alignment = LEFT
    r += 1

# ---------- Sheet 5: Components ----------
ws = wb.create_sheet("Components")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 26, 34, 60])
r = 1
r = section_banner(ws, r, "5. Component Architecture", 4, NAVY)
r = header_row(ws, r, ["#", "Component", "Module", "Responsibility"])
comps = [
    ("Ingestion", "services/ingestion.py", "Store uploaded PDF blob, create Transcript(status=uploaded), parse period_ordinal (FY*4+Q)."),
    ("Extraction", "services/extraction.py", "PDF bytes -> text (PyMuPDF default; pdfplumber/azure_docintel/stub). status=extracted."),
    ("Analysis", "services/analysis.py + llm/", "LLMClient -> TranscriptAnalysis: summary, sentiment, aggression, metrics/guidance/red_flags/section_scores."),
    ("LLM providers", "llm/client.py", "StubLLM (offline, deterministic), OpenAI/Azure (structured outputs), AgentFileLLM (real analysis, no key)."),
    ("Guidance reconciler", "services/guidance.py", "Match open promises vs new-quarter actuals -> MET/PARTIAL/MISSED/IN_PROGRESS + GuidanceOutcome (the differentiator)."),
    ("Scoring", "services/scoring.py", "10 section scores -> total_0_100 + growth/reliability/aggression/consistency + composite."),
    ("Fundamental ranking", "scripts/* + FundamentalRanking", "Business quality + Stock CAGR + Sales CAGR + ROE + PEG -> composite -> Rank 1..N."),
    ("Presentation", "app/static/*", "Vanilla-JS SPA (no build step); renders leaderboards, insights, PEG cards."),
]
for i, (nm, mod, resp) in enumerate(comps, 1):
    r = data_row(ws, r, [i, nm, mod, resp], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = BOLD_F
    ws.cell(row=r-1, column=3).font = MONO_F

# ---------- Sheet 6: Data Model ----------
ws = wb.create_sheet("Data Model")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 24, 30, 60])
r = 1
r = section_banner(ws, r, "6. Data Model — Entities", 4, NAVY)
r = header_row(ws, r, ["#", "Entity", "Keyed by", "Notes"])
ents = [
    ("Company", "id", "One row per stock (name, ticker)."),
    ("Transcript", "id, company_id", "One row per concall PDF; status uploaded->extracted->reconciled->scored; period_ordinal."),
    ("Metric", "transcript_id", "Extracted financial metric (revenue/ebitda/pat/margin)."),
    ("Guidance", "company_id + source_transcript_id", "Management commitment (NOT transcript_id)."),
    ("GuidanceOutcome", "guidance_id", "Whether guidance was met; the promise->outcome chain node."),
    ("RedFlag", "transcript_id", "Risk/concern detected in commentary."),
    ("SectionScore", "company_id + as_of_period_ordinal + section_no", "0-10 per template section."),
    ("CompanyScore", "company_id + as_of_period_ordinal", "Score col = total_0_100; also composite, growth_score, guidance_reliability_score."),
    ("ResearchSheet", "company_id", "Generated research output."),
    ("FundamentalRanking", "company_id", "rank, CAGRs, ROE + 5 PEG columns."),
]
for i, (e, k, n) in enumerate(ents, 1):
    r = data_row(ws, r, [i, e, k, n], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = BOLD_F
    ws.cell(row=r-1, column=3).font = MONO_F
r += 1
r = section_banner(ws, r, "Foreign-key contract (respect on deletes/joins)", 4, AMBER)
r = header_row(ws, r, ["Table", "Keyed by", "", ""], AMBER)
fk = [
    ("Metric, RedFlag", "transcript_id"),
    ("Guidance", "company_id + source_transcript_id"),
    ("GuidanceOutcome", "guidance_id"),
    ("SectionScore, CompanyScore", "company_id (+ as_of_period_ordinal)"),
]
for i, (t, k) in enumerate(fk, 1):
    ws.cell(row=r, column=1, value=t).font = BOLD_F
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=1).border = BORDER; ws.cell(row=r, column=1).fill = fill(LGREY if i % 2 == 0 else WHITE)
    cc = ws.cell(row=r, column=2, value=k); cc.font = MONO_F; cc.border = BORDER; cc.fill = fill(LGREY if i % 2 == 0 else WHITE); cc.alignment = LEFT
    for col in (3, 4):
        ws.cell(row=r, column=col).border = BORDER; ws.cell(row=r, column=col).fill = fill(LGREY if i % 2 == 0 else WHITE)
    r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws.cell(row=r, column=1, value="Temporal model: period_ordinal = FY*4 + Q gives a total order over quarters, powering "
           "guidance due-dates, as-of scoring snapshots, and multi-year chain assembly without date parsing at query time.")
c.font = BODY_F; c.alignment = WRAP; c.fill = fill(LBLUE2); c.border = BORDER
ws.row_dimensions[r].height = 44

# ---------- Sheet 7: Data Flows ----------
ws = wb.create_sheet("Data Flows")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 22, 78])
r = 1
r = section_banner(ws, r, "7.1 New Transcript Ingestion (incremental)", 3, NAVY)
r = header_row(ws, r, ["Step", "Stage", "Action"])
ing = [
    ("Ingestion", "Store blob + create Transcript(status=uploaded); parse period_ordinal."),
    ("Extraction", "PDF -> text; status=extracted."),
    ("Analysis", "LLM -> metrics / guidance / red_flags / section_scores."),
    ("Guidance", "Reconcile prior open promises vs this quarter's actuals; status=reconciled."),
    ("Scoring", "Recompute CompanyScore as-of this quarter; status=scored."),
    ("Dashboard", "That company's insights update automatically on next API read."),
]
for i, (s, a) in enumerate(ing, 1):
    r = data_row(ws, r, [i, s, a], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = BOLD_F
r += 1
r = section_banner(ws, r, "7.2 Ranking Refresh (offline batch)", 3, GREEN)
r = header_row(ws, r, ["Step", "Script", "Purpose"], GREEN)
ref = [
    ("build_fundamental_ranking.py", "Recompute composite + PEG factor."),
    ("persist_fundamental_ranking.py", "Write ranking rows to DB."),
    ("fetch_valuation_peg.py", "Refresh trailing PE + growth (if stale)."),
    ("persist_valuation_peg.py", "Write the 5 PEG columns."),
]
for i, (s, a) in enumerate(ref, 1):
    r = data_row(ws, r, [i, s, a], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = MONO_F

# ---------- Sheet 8: Valuation (PEG) ----------
ws = wb.create_sheet("Valuation (PEG)")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 30, 74])
r = 1
r = section_banner(ws, r, "8. Section 9 — PEG Valuation", 3, NAVY)
rows = [
    ("Formula", "PEG = trailing_PE / growth_pct"),
    ("Verdict", "PEG < 1 = Undervalued ; else Overvalued ; N/A if PE or growth missing."),
    ("Growth priority", "C = computed Profit CAGR  ->  S = sustainable Sales CAGR  ->  Y = Yahoo YoY earnings growth (volatile, flagged)."),
    ("Composite weight", "Valuation factor = 20% of composite; = inverted min-max of PEG capped at 4.0 (low PEG boosts)."),
    ("Missing data", "No PEG -> no boost, never a penalty. Loss-makers / no-growth = N/A (GOKEX, HFCL, RAIN, STLTECH, IDEAFORGE)."),
    ("Observed effect", "SANSERA #11->#4, SAMBHV #17->#8 (both PEG<1); ACUTAAS / NEOGEN dropped."),
]
r = header_row(ws, r, ["#", "Aspect", "Detail"])
for i, (k, v) in enumerate(rows, 1):
    r = data_row(ws, r, [i, k, v], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = BOLD_F
r += 1
r = section_banner(ws, r, "Composite scoring weights", 3, GREEN)
r = header_row(ws, r, ["Factor", "Weight", "Source"], GREEN)
wts = [
    ("Business Quality", "30%", "10-section template score"),
    ("Stock Price CAGR", "25%", "Yahoo chart API (winsorized 100%)"),
    ("Sales CAGR", "15%", "Transcript-mined (flags M/C/G)"),
    ("ROE", "10%", "Transcript-mined (P = ROCE-proxy)"),
    ("Valuation (PEG)", "20%", "PEG = PE / growth; low PEG -> boost"),
]
for i, (f, w, s) in enumerate(wts, 1):
    r = data_row(ws, r, [f, w, s], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=1).font = BOLD_F
    ws.cell(row=r-1, column=2).alignment = CENTER

# ---------- Sheet 9: Scalability ----------
ws = wb.create_sheet("Scalability")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 22, 34, 44])
r = 1
r = section_banner(ws, r, "9. Scalability & Deployment View", 4, NAVY)
r = header_row(ws, r, ["#", "Concern", "Prototype (today)", "Scale-out path"])
sc = [
    ("API", "Single uvicorn on :8010", "N replicas behind LB (stateless)"),
    ("DB", "SQLite file", "Postgres/Aurora (swap database_url)"),
    ("Pipeline", "In-process sequential", "Each stage -> queue consumer (Celery/SQS/Kafka)"),
    ("Blob storage", "data/blobs/ local", "S3 / Azure Blob (blob_dir abstraction)"),
    ("LLM", "Stub / agent-file (no key)", "OpenAI / Azure OpenAI structured outputs"),
    ("Extraction", "PyMuPDF", "Azure Document Intelligence for scanned PDFs"),
    ("Market data", "Offline scripts -> JSON -> DB", "Scheduled workers + caching layer"),
]
for i, (c1, c2, c3) in enumerate(sc, 1):
    r = data_row(ws, r, [i, c1, c2, c3], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = BOLD_F
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws.cell(row=r, column=1, value="Pipeline stages are already discrete idempotent functions, so promoting them to independent "
           "queue consumers is a mechanical change, not a redesign.")
c.font = Font(italic=True, color="404040"); c.alignment = WRAP; c.fill = fill(LAMBER); c.border = BORDER
ws.row_dimensions[r].height = 34

# ---------- Sheet 10: Decisions & Risks ----------
ws = wb.create_sheet("Decisions & Risks")
ws.sheet_view.showGridLines = False
set_widths(ws, [4, 40, 60])
r = 1
r = section_banner(ws, r, "10.1 Design Decisions & Trade-offs", 3, NAVY)
r = header_row(ws, r, ["#", "Decision", "Rationale"])
dec = [
    ("SQLite first", "Fastest path to a working, queryable memory; ORM keeps Postgres one line away."),
    ("Stub/agent LLM providers", "Entire pipeline runs with NO API key -> demoable, privacy-preserving; real providers plug in."),
    ("period_ordinal integer clock", "Simplifies temporal joins & guidance due-dating vs dates."),
    ("Offline ranking scripts", "Market data fetched in controlled batches and persisted; request path stays fast & resilient."),
    ("Sustainable CAGR over raw YoY for PEG", "Avoids false 'undervalued' signals from one-off spikes (NEOGEN 375%, SAMBHV 162%)."),
]
for i, (d, rr) in enumerate(dec, 1):
    r = data_row(ws, r, [i, d, rr], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = BOLD_F
r += 1
r = section_banner(ws, r, "10.2 Risks & Mitigations", 3, AMBER)
r = header_row(ws, r, ["#", "Risk", "Mitigation"], AMBER)
risks = [
    ("Transcript unavailable (recent IPO / only notices)", "Mark company blocked with reason (HBLENGINE, E2E)."),
    ("Third-party rate-limits / login-gates", "Offline batch fetch + persist; degrade to N/A, never fabricate."),
    ("LLM hallucination", "Structured-output schema + evidence quotes stored with every fact."),
    ("Inflated CAGR for <3y listings", "Flag *, winsorize at 100% in composite."),
    ("Data drift between quarters", "Idempotent pipeline re-run + as-of scoring snapshots."),
]
for i, (rk, mg) in enumerate(risks, 1):
    r = data_row(ws, r, [i, rk, mg], zebra=(i % 2 == 0))
    ws.cell(row=r-1, column=2).font = BOLD_F
r += 1
r = section_banner(ws, r, "10.3 Future Enhancements", 3, GREEN)
r = header_row(ws, r, ["#", "Enhancement", ""], GREEN)
fut = [
    "Promote pipeline stages to a real message queue (event-per-stage).",
    "Postgres migration + Alembic migrations (replace ad-hoc ALTER TABLE).",
    "Sortable/filterable ranking columns; sector-relative valuation percentiles.",
    "Real Screener 5-yr Sales/Profit CAGR + ROE ingestion.",
    "AuthN/AuthZ, multi-user watchlists, and audit trail.",
]
for i, f in enumerate(fut, 1):
    ws.cell(row=r, column=1, value=i).border = BORDER; ws.cell(row=r, column=1).fill = fill(LGREY if i % 2 == 0 else WHITE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cc = ws.cell(row=r, column=2, value=f); cc.font = BODY_F; cc.border = BORDER; cc.alignment = WRAP
    cc.fill = fill(LGREY if i % 2 == 0 else WHITE)
    ws.cell(row=r, column=3).border = BORDER; ws.cell(row=r, column=3).fill = fill(LGREY if i % 2 == 0 else WHITE)
    r += 1

# freeze header rows lightly on data sheets
for name in wb.sheetnames:
    wb[name].sheet_properties.tabColor = BLUE

wb.save(OUT)
print("SAVED", OUT)
print("Sheets:", wb.sheetnames)
